"""XLSX Extract tool."""

from datetime import datetime
from pathlib import Path
from typing import Any

from fastmcp import FastMCP


def _convert_cell_value(value: Any) -> Any:
    """Convert Excel cell values to JSON-serializable types."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def register_tools(mcp: FastMCP) -> None:
    """Register XLSX extraction tools with the MCP server."""

    @mcp.tool()
    def xlsx_read(
        path: str,
        sheet_name: str | None = None,
        start_row: int = 1,
        max_rows: int = 1000,
        max_cols: int | None = None,
    ) -> dict:
        """
        Extract data from an XLSX file safely with pagination and limits.

        Args:
            path: Path to the XLSX file
            sheet_name: Specific sheet to read. If None, reads the active sheet.
            start_row: First row to extract (1-indexed)
            max_rows: Maximum rows to extract in this call (default 1000).
                Use pagination if the sheet has more rows.
            max_cols: Maximum columns to extract per row. If None, extracts all.

        Returns:
            Dict containing rows, source metadata, and next_row token.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {"error": "openpyxl not installed."}

        file_path = Path(path).resolve()
        if not file_path.exists():
            return {"error": f"XLSX file not found: {path}"}
        if not file_path.is_file():
            return {"error": f"Not a file: {path}"}
        if file_path.suffix.lower() not in [".xlsx", ".xlsm"]:
            return {"error": f"Not an XLSX/XLSM file: {path}"}

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)

            try:
                if sheet_name:
                    if sheet_name not in wb.sheetnames:
                        return {
                            "error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
                        }
                    ws = wb[sheet_name]
                else:
                    ws = wb.active

                if ws is None:
                    return {"error": "Workbook has no active sheet"}

                total_rows = ws.max_row

                if total_rows is None:
                    # In some read_only cases max_row is None until iterated
                    # But we can iterate. We'll handle it dynamically.
                    pass
                else:
                    if start_row < 1:
                        start_row = 1
                    if total_rows > 0 and start_row > total_rows:
                        return {
                            "error": f"start_row {start_row} out of bounds "
                            f"(total rows: {total_rows})"
                        }

                rows_data = []
                rows_extracted = 0
                next_row = None

                # iter_rows is 1-indexed for min_row/max_row
                # Since max_row might be None in read_only mode, we can't reliably rely on it,
                # but we can pass min_row and manually stop at max_rows.

                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=start_row, values_only=True), start=start_row
                ):
                    if rows_extracted >= max_rows:
                        next_row = row_idx
                        break

                    row_values = []
                    for col_idx, cell in enumerate(row):
                        if max_cols is not None and col_idx >= max_cols:
                            break
                        row_values.append(_convert_cell_value(cell))

                    rows_data.append({"row_number": row_idx, "values": row_values})
                    rows_extracted += 1

                # If we didn't break out of the loop because we hit max_rows,
                # then we exhausted the sheet. next_row remains None.

                result = {
                    "rows": rows_data,
                    "metadata": {
                        "source": str(file_path.name),
                        "sheet": ws.title,
                        "rows_extracted": rows_extracted,
                    },
                }

                if next_row is not None:
                    result["next_row"] = next_row

                return result

            finally:
                wb.close()

        except Exception as e:
            return {"error": f"Failed to extract data from XLSX: {str(e)}"}
